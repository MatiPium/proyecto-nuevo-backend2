from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required, permission_required
from django.db import IntegrityError
from django.utils import timezone
from datetime import timedelta
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q, Count, Avg
from django.http import JsonResponse, HttpResponse
from django.core.exceptions import PermissionDenied

# Para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Importar modelos
from .models import Device, Measurement, Alert, DeviceType, DeviceStatus, AlertType
from accounts.models import Category, Zone, Organization  # ← AGREGAR ESTA LÍNEA
from .forms import DeviceForm, MeasurementForm, AlertForm, DeviceFilterForm


# -------------------------------------
# 🔐 Autenticación
# -------------------------------------

def login_view(request):
    if request.method == 'POST':
        identifier = (request.POST.get('email') or '').strip()   # puede ser email o username
        password   = (request.POST.get('password') or '')
        next_url   = request.POST.get('next') or request.GET.get('next')

        # Resolver a username si nos pasaron un email
        username = identifier
        if '@' in identifier:
            User = get_user_model()
            user_obj = User.objects.filter(email__iexact=identifier).first()
            username = user_obj.get_username() if user_obj else None

        user = authenticate(request, username=username, password=password) if username else None
        if user and user.is_active:
            login(request, user)
            return redirect(next_url or 'dashboard')  # cambia 'dashboard' si tu nombre de url es otro

        messages.error(request, 'Email o contraseña incorrectos')

    return render(request, 'devices/login.html')


def register_view(request):
    if request.method == 'POST':
        company_name = request.POST['company_name']
        email = request.POST['email']
        password = request.POST['password']
        password_confirm = request.POST['password_confirm']

        if password != password_confirm:
            return render(request, 'devices/register.html', {
                'error': 'Las contraseñas no coinciden'
            })

        if len(password) < 12:
            return render(request, 'devices/register.html', {
                'error': 'La contraseña debe tener al menos 12 caracteres'
            })

        try:
            user = User.objects.create_user(
                username=email,
                email=email,
                password=password,
                first_name=company_name
            )
            # 🔹 ELIMINA ESTA LÍNEA - Organization no existe
            # Organization.objects.create(name=company_name)

            return render(request, 'devices/register.html', {
                'success': f'¡Registro exitoso! La empresa {company_name} ha sido registrada correctamente.'
            })
        except IntegrityError:
            return render(request, 'devices/register.html', {
                'error': 'Este correo electrónico ya está registrado'
            })
        except Exception:
            return render(request, 'devices/register.html', {
                'error': 'Error al registrar la empresa. Intenta nuevamente.'
            })

    return render(request, 'devices/register.html')


def password_reset(request):
    message_sent = False
    if request.method == "POST":
        email = request.POST.get('email')
        message_sent = True
    return render(request, 'devices/password_reset.html', {'message_sent': message_sent})




def logout_view(request):
    logout(request)
    return redirect("login")


# ---------------------------
# 📊 Dashboard
# ---------------------------

@login_required
def dashboard(request):
    """Dashboard con estadísticas de la organización"""
    # Obtener organización del usuario
    user_org = None
    if hasattr(request.user, 'profile') and request.user.profile.organization:
        user_org = request.user.profile.organization
    
    # Filtrar datos por organización
    if user_org:
        devices = Device.objects.filter(zone__organization=user_org)
        measurements = Measurement.objects.filter(device__zone__organization=user_org)
        alerts = Alert.objects.filter(device__zone__organization=user_org)
    else:
        devices = Device.objects.filter(owner=request.user)
        measurements = Measurement.objects.filter(device__owner=request.user)
        alerts = Alert.objects.filter(device__owner=request.user)
    
    # Estadísticas
    total_devices = devices.count()
    active_devices = devices.filter(status='active').count()  # ← CAMBIADO de is_active=True a status='active'
    inactive_devices = devices.filter(status='inactive').count()
    maintenance_devices = devices.filter(status='maintenance').count()
    
    total_measurements = measurements.count()
    unresolved_alerts = alerts.filter(is_resolved=False).count()
    
    # Últimas mediciones
    recent_measurements = measurements.select_related('device').order_by('-timestamp')[:10]
    
    # Alertas recientes no resueltas
    recent_alerts = alerts.filter(is_resolved=False).select_related('device').order_by('-created_at')[:5]
    
    # Dispositivos por tipo
    devices_by_type = {}
    for choice in Device._meta.get_field('device_type').choices:
        type_code = choice[0]
        type_name = choice[1]
        count = devices.filter(device_type=type_code).count()
        if count > 0:
            devices_by_type[type_name] = count
    
    # Dispositivos por estado
    devices_by_status = {}
    for choice in Device._meta.get_field('status').choices:
        status_code = choice[0]
        status_name = choice[1]
        count = devices.filter(status=status_code).count()
        if count > 0:
            devices_by_status[status_name] = count
    
    # Alertas por tipo
    alerts_by_type = {}
    for choice in Alert._meta.get_field('alert_type').choices:
        type_code = choice[0]
        type_name = choice[1]
        count = alerts.filter(alert_type=type_code, is_resolved=False).count()
        if count > 0:
            alerts_by_type[type_name] = count
    
    context = {
        'total_devices': total_devices,
        'active_devices': active_devices,
        'inactive_devices': inactive_devices,
        'maintenance_devices': maintenance_devices,
        'total_measurements': total_measurements,
        'unresolved_alerts': unresolved_alerts,
        'recent_measurements': recent_measurements,
        'recent_alerts': recent_alerts,
        'devices_by_type': devices_by_type,
        'devices_by_status': devices_by_status,
        'alerts_by_type': alerts_by_type,
        'user_org': user_org,
    }
    
    return render(request, 'devices/dashboard.html', context)


# ---------------------------
# 💻 Dispositivos
# ---------------------------

@login_required
@permission_required('devices.view_device', raise_exception=True)
def device_list(request):
    """Lista dispositivos de la misma organización"""
    # Obtener la organización del usuario
    user_org = None
    if hasattr(request.user, 'profile') and request.user.profile.organization:
        user_org = request.user.profile.organization
    
    # Filtrar dispositivos por organización
    if user_org:
        devices = Device.objects.filter(
            zone__organization=user_org
        ).select_related('category', 'zone', 'owner')  # ← ELIMINADO 'device_type'
    else:
        # Si no tiene organización, solo ve sus propios dispositivos
        devices = Device.objects.filter(owner=request.user).select_related('category', 'zone', 'owner')
    
    # Filtros de búsqueda
    search = request.GET.get('search', '')
    device_type = request.GET.get('device_type', '')
    category = request.GET.get('category', '')
    status = request.GET.get('status', '')
    zone = request.GET.get('zone', '')
    
    if search:
        devices = devices.filter(
            Q(name__icontains=search) | 
            Q(description__icontains=search)
        )
    if device_type:
        devices = devices.filter(device_type=device_type)
    if category:
        devices = devices.filter(category_id=category)
    if status:
        devices = devices.filter(status=status)
    if zone:
        devices = devices.filter(zone_id=zone)
    
    # Agregar permisos de edición
    for device in devices:
        device.can_edit = (device.owner == request.user) or request.user.is_superuser
        device.can_delete = (device.owner == request.user) or request.user.is_superuser
    
    # Datos para filtros
    categories = Category.objects.all()
    zones = Zone.objects.filter(organization=user_org) if user_org else Zone.objects.all()
    
    # Obtener choices de device_type del modelo
    from .models import DeviceType
    device_types = DeviceType.choices
    
    # Obtener choices de status del modelo
    from .models import DeviceStatus
    status_choices = DeviceStatus.choices
    
    context = {
        'devices': devices,
        'categories': categories,
        'zones': zones,
        'device_types': device_types,
        'status_choices': status_choices,
        'search': search,
        'selected_type': device_type,
        'selected_category': category,
        'selected_status': status,
        'selected_zone': zone,
        'user_org': user_org,
    }
    
    return render(request, 'devices/device_list.html', context)

@login_required
@permission_required('devices.view_device', raise_exception=True)
def device_detail(request, pk):
    """Detalle de un dispositivo"""
    device = get_object_or_404(Device, pk=pk, owner=request.user)
    
    # Últimas mediciones del dispositivo
    measurements = device.measurements.all().order_by('-timestamp')[:20]
    
    # Alertas del dispositivo
    alerts = device.alerts.all().order_by('-created_at')[:10]
    
    context = {
        'device': device,
        'measurements': measurements,
        'alerts': alerts,
        'can_edit': request.user.has_perm('devices.change_device'),
        'can_delete': request.user.has_perm('devices.delete_device'),
    }
    return render(request, 'devices/device_detail.html', context)


@login_required
@permission_required('devices.add_device', raise_exception=True)
def create_device(request):
    """Crear un nuevo dispositivo"""
    if request.method == 'POST':
        form = DeviceForm(request.POST)
        if form.is_valid():
            device = form.save(commit=False)
            device.owner = request.user
            device.save()
            messages.success(request, f'Dispositivo "{device.name}" creado exitosamente.')
            return redirect('device_list')
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = DeviceForm()
    
    # Obtener todos los tipos de dispositivos
    device_types = DeviceType.objects.all()
    
    context = {
        'form': form,
        'device_types': device_types,
    }
    return render(request, 'devices/device_form.html', context)


@login_required
@permission_required('devices.change_device', raise_exception=True)
def update_device(request, pk):
    """Actualizar un dispositivo existente"""
    device = get_object_or_404(Device, pk=pk, owner=request.user)

    if request.method == 'POST':
        form = DeviceForm(request.POST, instance=device)
        if form.is_valid():
            device = form.save()
            messages.success(request, f'Dispositivo "{device.name}" actualizado exitosamente.')      
            return redirect('device_list')
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = DeviceForm(instance=device)

    device_types = DeviceType.objects.all()

    context = {
        'form': form,
        'device': device,
        'device_types': device_types,
    }
    return render(request, 'devices/update_device.html', context)  # Usar update_device.html


@login_required
@permission_required('devices.delete_device', raise_exception=True)
def delete_device(request, pk):
    """Eliminar dispositivo"""
    device = get_object_or_404(Device, pk=pk, owner=request.user)
    
    if request.method == 'POST':
        device_name = device.name
        device.delete()
        messages.success(request, f'Dispositivo "{device_name}" eliminado exitosamente.')
        return redirect('device_list')
    
    return render(request, 'devices/device_confirm_delete.html', {
        'device': device
    })


@login_required
@permission_required('devices.view_measurement', raise_exception=True)
def measurement_list(request):
    """Lista de mediciones con búsqueda, ordenamiento y paginación"""
    
    # Obtener mediciones del usuario
    measurements = Measurement.objects.filter(
        device__owner=request.user
    ).select_related('device', 'device__device_type')
    
    # 🔍 BÚSQUEDA
    search_query = request.GET.get('q', '').strip()
    if search_query:
        measurements = measurements.filter(
            Q(device__name__icontains=search_query) |
            Q(unit__icontains=search_query) |
            Q(value__icontains=search_query)
        )
    
    # 📊 FILTROS ADICIONALES
    device_id = request.GET.get('device')
    if device_id:
        measurements = measurements.filter(device_id=device_id)
    
    date_from = request.GET.get('date_from')
    if date_from:
        measurements = measurements.filter(timestamp__gte=date_from)
    
    date_to = request.GET.get('date_to')
    if date_to:
        measurements = measurements.filter(timestamp__lte=date_to)
    
    # 📊 ORDENAMIENTO
    sort = request.GET.get('sort', '-timestamp')
    valid_sort_fields = [
        'timestamp', '-timestamp',
        'value', '-value',
        'device__name', '-device__name',
        'unit', '-unit'
    ]
    
    if sort in valid_sort_fields:
        measurements = measurements.order_by(sort)
    else:
        measurements = measurements.order_by('-timestamp')
    
    # 📄 PAGINACIÓN
    page_size = request.GET.get('page_size')
    
    if page_size:
        try:
            page_size = int(page_size)
            request.session['page_size_measurements'] = page_size
        except ValueError:
            page_size = request.session.get('page_size_measurements', 15)
    else:
        page_size = request.session.get('page_size_measurements', 15)
    
    if page_size not in [5, 10, 15, 25, 50, 100]:
        page_size = 15
    
    paginator = Paginator(measurements, page_size)
    page_number = request.GET.get('page', 1)
    
    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.get_page(1)
    except EmptyPage:
        page_obj = paginator.get_page(paginator.num_pages)
    
    # Dispositivos para el filtro
    user_devices = Device.objects.filter(owner=request.user, is_active=True)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'current_sort': sort,
        'page_size': page_size,
        'total_measurements': paginator.count,
        'user_devices': user_devices,
        'selected_device': device_id,
        'date_from': date_from,
        'date_to': date_to,
        'can_add': request.user.has_perm('devices.add_measurement'),
    }
    return render(request, 'devices/measurement_list.html', context)


@login_required
@permission_required('devices.add_measurement', raise_exception=True)
def create_measurement(request):
    """Crear una nueva medición"""
    if request.method == 'POST':
        form = MeasurementForm(request.POST)
        
        # Validar que el dispositivo pertenezca al usuario
        device_id = request.POST.get('device')
        if device_id:
            try:
                device = Device.objects.get(id=device_id, owner=request.user)
                if form.is_valid():
                    measurement = form.save()
                    messages.success(request, f'Medición registrada exitosamente para {device.name}.')
                    return redirect('measurement_list')
            except Device.DoesNotExist:
                messages.error(request, 'El dispositivo seleccionado no existe o no tienes permisos.')
        else:
            messages.error(request, 'Debes seleccionar un dispositivo.')
    else:
        form = MeasurementForm()
    
    # Obtener solo los dispositivos del usuario actual
    user_devices = Device.objects.filter(owner=request.user, is_active=True)
    
    context = {
        'form': form,
        'user_devices': user_devices,
    }
    return render(request, 'devices/measurement_form.html', context)


@login_required
@permission_required('devices.change_measurement', raise_exception=True)
def edit_measurement(request, pk):
    """Editar una medición existente"""
    measurement = get_object_or_404(Measurement, pk=pk)
    
    # Verificar que el usuario sea dueño del dispositivo
    if measurement.device.owner != request.user and not request.user.is_staff:
        messages.error(request, 'No tienes permiso para editar esta medición.')
        return redirect('measurement_list')
    
    if request.method == 'POST':
        form = MeasurementForm(request.POST, instance=measurement)
        if form.is_valid():
            form.save()
            messages.success(request, 'Medición actualizada exitosamente.')
            return redirect('measurement_list')
    else:
        form = MeasurementForm(instance=measurement)
    
    user_devices = Device.objects.filter(owner=request.user, is_active=True)
    
    context = {
        'form': form,
        'user_devices': user_devices,
        'measurement': measurement,
    }
    return render(request, 'devices/measurement_form.html', context)


@login_required
@permission_required('devices.delete_measurement', raise_exception=True)
def delete_measurement(request, pk):
    """Eliminar una medición"""
    measurement = get_object_or_404(Measurement, pk=pk)
    
    # Verificar que el usuario sea dueño del dispositivo
    if measurement.device.owner != request.user and not request.user.is_staff:
        messages.error(request, 'No tienes permiso para eliminar esta medición.')
        return redirect('measurement_list')
    
    if request.method == 'POST':
        device_name = measurement.device.name
        measurement.delete()
        messages.success(request, f'Medición del dispositivo {device_name} eliminada exitosamente.')
        return redirect('measurement_list')
    
    context = {
        'measurement': measurement,
    }
    return render(request, 'devices/measurement_confirm_delete.html', context)


@login_required
@permission_required('devices.view_alert', raise_exception=True)
def alert_list(request):
    """Lista de alertas"""
    alerts = Alert.objects.filter(
        device__owner=request.user
    ).select_related('device').order_by('-created_at')
    
    # Filtrar por estado
    status = request.GET.get('status', 'all')
    if status == 'active':
        alerts = alerts.filter(is_resolved=False)
    elif status == 'resolved':
        alerts = alerts.filter(is_resolved=True)
    
    context = {
        'alerts': alerts,
        'status': status,
    }
    return render(request, 'devices/alert_list.html', context)


@login_required
@permission_required('devices.view_device', raise_exception=True)
def export_devices_excel(request):
    """Exportar dispositivos a Excel"""
    
    # Obtener el mismo queryset que en device_list (con filtros)
    devices = Device.objects.filter(owner=request.user).select_related('device_type')
    
    # Aplicar búsqueda si existe
    search_query = request.GET.get('q', '').strip()
    if search_query:
        devices = devices.filter(
            Q(name__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(device_type__name__icontains=search_query)
        )
    
    # Aplicar ordenamiento si existe
    sort = request.GET.get('sort', '-created_at')
    valid_sort_fields = [
        'name', '-name',
        'device_type__name', '-device_type__name',
        'location', '-location',
        'is_active', '-is_active',
        'created_at', '-created_at'
    ]
    
    if sort in valid_sort_fields:
        devices = devices.order_by(sort)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Dispositivos"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['ID', 'Nombre', 'Tipo de Dispositivo', 'Ubicación', 'Estado', 'Fecha de Creación', 'Última Actualización']
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Escribir datos
    for device in devices:
        row = [
            device.id,
            device.name,
            str(device.device_type),
            device.location or 'N/A',
            'Activo' if device.is_active else 'Inactivo',
            device.created_at.strftime('%d/%m/%Y %H:%M:%S'),
            device.updated_at.strftime('%d/%m/%Y %H:%M:%S')
        ]
        ws.append(row)
        
        # Aplicar bordes a las celdas de datos
        row_num = ws.max_row
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            if col_num == 5:  # Columna Estado
                if device.is_active:
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    
    # Ajustar ancho de columnas
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'dispositivos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Guardar workbook en la respuesta
    wb.save(response)
    
    return response


@login_required
@permission_required('devices.view_measurement', raise_exception=True)
def export_measurements_excel(request):
    """Exportar mediciones a Excel"""
    
    # Obtener el mismo queryset que en measurement_list (con filtros)
    measurements = Measurement.objects.filter(
        device__owner=request.user
    ).select_related('device', 'device__device_type')
    
    # Aplicar búsqueda
    search_query = request.GET.get('q', '').strip()
    if search_query:
        measurements = measurements.filter(
            Q(device__name__icontains=search_query) |
            Q(unit__icontains=search_query) |
            Q(value__icontains=search_query)
        )
    
    # Filtros adicionales
    device_id = request.GET.get('device')
    if device_id:
        measurements = measurements.filter(device_id=device_id)
    
    date_from = request.GET.get('date_from')
    if date_from:
        measurements = measurements.filter(timestamp__gte=date_from)
    
    date_to = request.GET.get('date_to')
    if date_to:
        measurements = measurements.filter(timestamp__lte=date_to)
    
    # Ordenamiento
    sort = request.GET.get('sort', '-timestamp')
    valid_sort_fields = [
        'timestamp', '-timestamp',
        'value', '-value',
        'device__name', '-device__name',
        'unit', '-unit'
    ]
    
    if sort in valid_sort_fields:
        measurements = measurements.order_by(sort)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Mediciones"
    
    # Estilos
    header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['ID', 'Dispositivo', 'Tipo', 'Valor', 'Unidad', 'Fecha y Hora']
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Escribir datos
    for measurement in measurements:
        row = [
            measurement.id,
            measurement.device.name,
            str(measurement.device.device_type),
            float(measurement.value),
            measurement.unit,
            measurement.timestamp.strftime('%d/%m/%Y %H:%M:%S')
        ]
        ws.append(row)
        
        # Aplicar bordes
        row_num = ws.max_row
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            
            # Formato para números
            if col_num == 4:  # Columna Valor
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
    
    # Ajustar ancho de columnas
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Agregar hoja de resumen
    ws_summary = wb.create_sheet("Resumen")
    ws_summary.append(['Estadísticas de Mediciones'])
    ws_summary.append([])
    ws_summary.append(['Total de mediciones:', measurements.count()])
    
    if measurements.exists():
        total_value = sum(float(m.value) for m in measurements)
        avg_value = total_value / measurements.count()
        max_value = max(float(m.value) for m in measurements)
        min_value = min(float(m.value) for m in measurements)
        
        ws_summary.append(['Valor total:', f'{total_value:.2f}'])
        ws_summary.append(['Valor promedio:', f'{avg_value:.2f}'])
        ws_summary.append(['Valor máximo:', f'{max_value:.2f}'])
        ws_summary.append(['Valor mínimo:', f'{min_value:.2f}'])
    
    # Estilo para la hoja de resumen
    ws_summary.cell(1, 1).font = Font(bold=True, size=14)
    for row in range(3, 8):
        ws_summary.cell(row, 1).font = Font(bold=True)
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'mediciones_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Guardar workbook en la respuesta
    wb.save(response)
    
    return response


# Handlers de error
def page_not_found(request, exception):
    """Vista personalizada para error 404"""
    return render(request, '404.html', status=404)


def permission_denied(request, exception):
    """Vista personalizada para error 403"""
    return render(request, '403.html', status=403)


def server_error(request):
    """Vista personalizada para error 500"""
    return render(request, '500.html', status=500)